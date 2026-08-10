from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
import logging
import os
import shutil
from pathlib import Path
from uuid import uuid4

import openpyxl
import pandas as pd

from ..db.database import get_db
from ..core.settings import settings
from ..db.models import PipelineJob, SampleStage, User
from ..services.job_service import audit, create_job
from ..tasks.pipeline_tasks import enqueue_pipeline_job
from ..utils import get_current_user
from ..utils_paths import ensure_safe_component, safe_resolve_user_path

router = APIRouter()
logger = logging.getLogger(__name__)

@router.post("/deg/run", status_code=202)
async def run_deg(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = await request.json()
    user_id = current_user.id
    contrast_ids = data.get("contrast_ids", [])
    genome_accession = data.get("genome_accession")
    logger.info(f"Iniciando DEG para user_id={user_id} com contrastes {contrast_ids} e genoma {genome_accession}")

    preprocess_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../users", str(user_id), "preprocess"))
    deg_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../users", str(user_id), "DEG"))
    os.makedirs(deg_dir, exist_ok=True)

    if not os.path.exists(preprocess_dir):
        logger.error(f"Diretório preprocess não encontrado: {preprocess_dir}")
        raise HTTPException(status_code=500, detail=f"Diretório preprocess não encontrado: {preprocess_dir}")

    # Buscar todos os contrastes do usuário e salvar contrasts_db.txt
    contrasts = db.query(SampleStage).filter(
        SampleStage.stage_id == 8,
        SampleStage.user_id == user_id,
        SampleStage.status == "Contrast"
    ).all()
    contrasts_db_path = os.path.join(preprocess_dir, "contrasts_db.txt")
    with open(contrasts_db_path, "w", encoding="utf-8") as f:
        f.write("id\tname\n")
        for c in contrasts:
            f.write(f"{c.id}\t{c.name}\n")

    # Salvar os contrastes selecionados em um arquivo para o script R ler
    selected_contrasts_path = os.path.join(preprocess_dir, "selected_contrasts.txt")
    with open(selected_contrasts_path, "w", encoding="utf-8") as f:
        for cid in contrast_ids:
            f.write(str(cid) + "\n")

    # Salvar o accession do genoma selecionado
    if genome_accession:
        genome_file_path = os.path.join(preprocess_dir, "selected_genome.txt")
        with open(genome_file_path, "w", encoding="utf-8") as f:
            f.write(str(genome_accession) + "\n")

    job = create_job(
        db,
        stage="deg",
        user_id=user_id,
        payload={"contrast_ids": contrast_ids, "genome_accession": genome_accession},
    )
    audit(
        db,
        action="deg_enqueued",
        user_id=user_id,
        stage="deg",
        job_id=job.id,
        metadata_json={"contrast_ids": contrast_ids, "genome_accession": genome_accession},
    )
    enqueue_pipeline_job(job.id)
    return {"job_id": job.id, "status": "PENDING", "message": "DEG job enqueued"}




# Endpoint incremental de progresso do DEG
@router.get("/pipeline/jobs/{job_id}/progress")
@router.get("/deg/jobs/{job_id}/progress")
async def get_deg_job_progress(
    job_id: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Combina o progresso geral do worker com as etapas detalhadas do DEG.R.

    O cursor possui dois componentes: linhas do job e linhas do R.
    """
    from ..db.models import PipelineJob

    job = db.query(PipelineJob).filter(
        PipelineJob.id == job_id,
        PipelineJob.user_id == current_user.id,
        PipelineJob.stage.in_(["deg", "clustering"]),
    ).first()

    if job is None:
        raise HTTPException(
            status_code=404,
            detail="Job DEG não encontrado.",
        )

    raw_cursor = request.query_params.get("cursor", "0:0")

    try:
        if ":" in raw_cursor:
            job_cursor_text, r_cursor_text = raw_cursor.split(
                ":",
                1,
            )
            job_cursor = max(0, int(job_cursor_text))
            r_cursor = max(0, int(r_cursor_text))
        else:
            # Compatibilidade com o cursor inteiro da versão anterior.
            job_cursor = 0
            r_cursor = max(0, int(raw_cursor))
    except (TypeError, ValueError):
        job_cursor = 0
        r_cursor = 0

    user_root = (
        Path(__file__).resolve().parents[3]
        / "users"
        / str(current_user.id)
    )

    job_log_path = (
        user_root
        / "logs"
        / "jobs"
        / f"{job_id}.progress.log"
    )
    r_log_path = user_root / "preprocess" / "DEG_R.log"

    def read_lines(path: Path) -> list[str]:
        if not path.is_file():
            return []

        try:
            return path.read_text(
                encoding="utf-8",
                errors="replace",
            ).splitlines()
        except OSError:
            return []

    job_lines = read_lines(job_log_path)

    # Não retorna DEG_R.log antigo enquanto o job atual não começou.
    r_lines = []

    if job.stage == "deg" and r_log_path.is_file():
        is_current = True

        if job.started_at is not None:
            try:
                is_current = (
                    r_log_path.stat().st_mtime
                    >= job.started_at.timestamp() - 2
                )
            except (OSError, ValueError):
                is_current = False

        if is_current:
            r_lines = read_lines(r_log_path)

    job_cursor = min(job_cursor, len(job_lines))
    r_cursor = min(r_cursor, len(r_lines))

    new_lines = (
        job_lines[job_cursor:]
        + r_lines[r_cursor:]
    )

    # Remove apenas duplicatas consecutivas.
    lines = []
    previous = None

    for line in new_lines:
        normalized = line.strip()

        if not normalized or normalized == previous:
            continue

        lines.append(normalized)
        previous = normalized

    return {
        "job_id": job.id,
        "status": job.status,
        "cursor": f"{len(job_lines)}:{len(r_lines)}",
        "lines": lines,
    }


def _xlsx_sheet_names(workbook_path: Path) -> list[str]:
    """Lê os nomes das abas sem carregar desenhos, gráficos ou imagens."""
    import zipfile
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(workbook_path) as archive:
        workbook_xml = archive.read("xl/workbook.xml")

    root = ET.fromstring(workbook_xml)
    namespace = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    }

    return [
        sheet.attrib["name"]
        for sheet in root.findall("main:sheets/main:sheet", namespace)
    ]


def _create_sanitized_workbook_copy(
    workbook_path: Path,
    temporary_path: Path,
) -> None:
    """
    Cria uma cópia OOXML sem drawings/charts e sem relacionamentos para
    desenhos. Isso permite recuperar planilhas com drawing*.xml ausente.
    """
    import zipfile
    import xml.etree.ElementTree as ET

    relationship_ns = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    worksheet_ns = (
        "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    )

    with zipfile.ZipFile(workbook_path, "r") as source:
        with zipfile.ZipFile(
            temporary_path,
            "w",
            compression=zipfile.ZIP_DEFLATED,
        ) as target:
            for item in source.infolist():
                name = item.filename

                if name.startswith((
                    "xl/drawings/",
                    "xl/charts/",
                    "xl/media/",
                )):
                    continue

                data = source.read(name)

                if (
                    name.startswith("xl/worksheets/")
                    and name.endswith(".xml")
                    and "/_rels/" not in name
                ):
                    root = ET.fromstring(data)

                    for tag_name in ("drawing", "legacyDrawing"):
                        for element in list(
                            root.findall(f"{{{worksheet_ns}}}{tag_name}")
                        ):
                            root.remove(element)

                    data = ET.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                    )

                elif (
                    name.startswith("xl/worksheets/_rels/")
                    and name.endswith(".rels")
                ):
                    root = ET.fromstring(data)

                    for relationship in list(root):
                        relationship_type = relationship.attrib.get(
                            "Type",
                            "",
                        )
                        if relationship_type.endswith((
                            "/drawing",
                            "/vmlDrawing",
                            "/image",
                            "/chart",
                        )):
                            root.remove(relationship)

                    data = ET.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                    )

                target.writestr(item, data)


def _remove_workbook_sheets(
    workbook_path: Path,
    requested_sheets: list[str],
) -> tuple[list[str], list[str]]:
    """
    Exclui abas reconstruindo uma planilha limpa.

    A reconstrução remove referências quebradas de desenhos e mantém cada
    célula, fórmula e aba restante. A substituição final é atômica.
    """
    import shutil
    import tempfile
    from uuid import uuid4

    if not workbook_path.exists():
        return [], []

    requested = set(requested_sheets)

    try:
        original_names = _xlsx_sheet_names(workbook_path)
    except Exception as exc:
        recovery_dir = workbook_path.parent / "recovery" / "corrupt"
        recovery_dir.mkdir(parents=True, exist_ok=True)

        recovered_path = recovery_dir / (
            f"{workbook_path.stem}-{uuid4().hex}.xlsx"
        )
        os.replace(workbook_path, recovered_path)

        logger.warning(
            "Planilha ilegível movida para recuperação: %s",
            recovered_path,
            exc_info=True,
        )

        # O arquivo ilegível deixa de bloquear novas análises.
        return list(requested_sheets), []

    targets = [
        name for name in original_names
        if name in requested
    ]
    remaining = [
        name for name in original_names
        if name not in requested
    ]

    if not targets:
        return [], original_names

    if not remaining:
        workbook_path.unlink(missing_ok=True)
        return targets, []

    sanitized_fd, sanitized_name = tempfile.mkstemp(
        prefix=".sanitized-",
        suffix=".xlsx",
        dir=workbook_path.parent,
    )
    os.close(sanitized_fd)
    sanitized_path = Path(sanitized_name)

    output_path = workbook_path.with_name(
        f".{workbook_path.stem}.{uuid4().hex}.xlsx"
    )
    rollback_path = workbook_path.with_name(
        f".{workbook_path.stem}.rollback.{uuid4().hex}.xlsx"
    )

    source_workbook = None
    output_workbook = None

    try:
        # Primeiro tenta leitura direta em modo somente leitura.
        try:
            source_workbook = openpyxl.load_workbook(
                workbook_path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except (KeyError, ValueError):
            _create_sanitized_workbook_copy(
                workbook_path,
                sanitized_path,
            )
            source_workbook = openpyxl.load_workbook(
                sanitized_path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )

        output_workbook = openpyxl.Workbook(write_only=True)

        for sheet_name in remaining:
            source_sheet = source_workbook[sheet_name]
            output_sheet = output_workbook.create_sheet(
                title=sheet_name
            )

            for row in source_sheet.iter_rows(values_only=True):
                output_sheet.append(row)

        output_workbook.save(output_path)
        source_workbook.close()

        # Confirma que o XLSX reconstruído abre antes de substituir o original.
        validation = openpyxl.load_workbook(
            output_path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        validated_names = list(validation.sheetnames)
        validation.close()

        if validated_names != remaining:
            raise ValueError(
                "As abas reconstruídas não correspondem às esperadas: "
                f"{validated_names} != {remaining}"
            )

        # Troca atômica com rollback.
        os.replace(workbook_path, rollback_path)

        try:
            os.replace(output_path, workbook_path)
        except Exception:
            os.replace(rollback_path, workbook_path)
            raise
        else:
            rollback_path.unlink(missing_ok=True)

        return targets, remaining

    finally:
        if source_workbook is not None:
            try:
                source_workbook.close()
            except Exception:
                pass

        sanitized_path.unlink(missing_ok=True)
        output_path.unlink(missing_ok=True)


@router.delete("/deg/sheets")
async def delete_deg_sheets(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exclui abas DEG e os resultados derivados correspondentes."""
    import shutil

    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Corpo JSON inválido.",
        )

    sheets = payload.get("sheets", [])

    if not isinstance(sheets, list):
        raise HTTPException(
            status_code=400,
            detail="'sheets' deve ser uma lista.",
        )

    sheets = list(dict.fromkeys(
        str(sheet).strip()
        for sheet in sheets
        if isinstance(sheet, str) and sheet.strip()
    ))

    if not sheets:
        raise HTTPException(
            status_code=400,
            detail="Selecione pelo menos uma aba para excluir.",
        )

    user_id = current_user.id
    user_dir = Path(__file__).resolve().parents[3] / "users" / str(user_id)
    deg_dir = user_dir / "DEG"

    workbook_paths = [
        deg_dir / "DEG.xlsx",
        deg_dir / "DEG_full.xlsx",
    ]

    if not any(path.is_file() for path in workbook_paths):
        raise HTTPException(
            status_code=404,
            detail="Nenhuma planilha DEG foi encontrada.",
        )

    removed = set()
    remaining_by_workbook = {}

    try:
        for workbook_path in workbook_paths:
            deleted, remaining = _remove_workbook_sheets(
                workbook_path,
                sheets,
            )
            removed.update(deleted)
            remaining_by_workbook[workbook_path.name] = remaining

        if not removed:
            raise HTTPException(
                status_code=404,
                detail="Nenhuma das abas selecionadas foi encontrada.",
            )

        # Remove apenas PNGs associados às abas excluídas.
        if deg_dir.is_dir():
            for image_path in deg_dir.glob("*.png"):
                if any(
                    image_path.stem.endswith(f" - {sheet_name}")
                    for sheet_name in removed
                ):
                    image_path.unlink(missing_ok=True)

        # Remove resultados derivados por contraste.
        for sheet_name in removed:
            for parent_name in ("clustering", "llm"):
                parent = (user_dir / parent_name).resolve()
                target = (parent / sheet_name).resolve()

                if (
                    target != parent
                    and parent in target.parents
                    and target.is_dir()
                ):
                    shutil.rmtree(target)

        # A limpeza auxiliar do banco não invalida a exclusão dos arquivos.
        try:
            db.query(SampleStage).filter(
                SampleStage.user_id == user_id,
                SampleStage.stage_id.in_([9, 10]),
                SampleStage.name.in_(list(removed)),
            ).delete(synchronize_session=False)
            db.commit()
        except Exception:
            db.rollback()
            logger.warning(
                "Não foi possível limpar registros derivados de %s",
                sorted(removed),
                exc_info=True,
            )

        logger.info(
            "Usuário %s excluiu abas DEG: %s",
            user_id,
            sorted(removed),
        )

        return {
            "message": "Abas excluídas com sucesso.",
            "deleted_sheets": sorted(removed),
            "remaining_sheets": remaining_by_workbook,
        }

    except HTTPException:
        raise
    except PermissionError:
        raise HTTPException(
            status_code=409,
            detail="Uma planilha DEG está em uso. Tente novamente.",
        )
    except Exception as exc:
        logger.exception("Erro ao excluir abas DEG: %s", exc)
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao excluir abas: {exc}",
        )


@router.get("/deg/sheets")
async def get_deg_sheets(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user_id = current_user.id
    deg_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../users", str(user_id), "DEG"))
    deg_xlsx = os.path.join(deg_dir, "DEG.xlsx")
    if not os.path.exists(deg_xlsx):
        raise HTTPException(status_code=404, detail="Arquivo DEG.xlsx não encontrado.")
    try:
        sheet_names = _xlsx_sheet_names(Path(deg_xlsx))
        return {"sheets": sheet_names}
    except Exception as e:
        logger.error(f"Erro ao ler abas do DEG.xlsx: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao ler abas do DEG.xlsx: {e}")

@router.get("/deg/sheet_data")
async def get_deg_sheet_data(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user_id = current_user.id
    sheet_name = request.query_params.get("sheet")
    page = max(1, int(request.query_params.get("page", "1")))
    page_size = min(500, max(1, int(request.query_params.get("page_size", "100"))))
    logger.info(f"[DEG] Requisição para sheet_data: user_id={user_id}, sheet_name={sheet_name}")
    
    if not sheet_name:
        raise HTTPException(status_code=400, detail="Sheet name is required.")
    
    deg_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../users", str(user_id), "DEG"))
    deg_xlsx = os.path.join(deg_dir, "DEG.xlsx")
    logger.info(f"[DEG] Caminho do arquivo DEG.xlsx: {deg_xlsx}")
    
    if not os.path.exists(deg_xlsx):
        raise HTTPException(status_code=404, detail="Arquivo DEG.xlsx não encontrado.")
    
    try:
        df = pd.read_excel(deg_xlsx, sheet_name=sheet_name)
        if df.empty:
            return {"columns": [], "rows": [], "page": page, "page_size": page_size, "total_rows": 0}

        df = df.fillna("")

        total_rows = len(df)
        start = (page - 1) * page_size
        end = start + page_size
        paged_df = df.iloc[start:end]

        columns = df.columns.tolist()
        rows = paged_df.values.tolist()

        return JSONResponse(content={
            "columns": columns,
            "rows": rows,
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
        })
        
    except ValueError as e:
        if "Worksheet named" in str(e):
            raise HTTPException(status_code=404, detail=f"Aba '{sheet_name}' não encontrada no arquivo.")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"[DEG] Erro ao processar o arquivo Excel com pandas: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao processar o arquivo Excel: {e}")
